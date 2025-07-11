import pandas as pd
import requests
from io import StringIO

from .config import drop_cols, cleaning_dict, translation_columns, translation_answers
from .validation import gender_age, valid, attention_col
from .dac_analysis import set_dependent_questions, set_no_knowledge_to_neutral

survey_id1 = 1740754  # Direct Air Capture in Germany
survey_id2 = 1837044  # Direct Air Capture in Germany - Bilendi
api_url = "https://api.surveyhero.com/v1/surveys/{survey_id}/responses"

survey_id_bilendi = survey_id2

all_surveys = [survey_id1, survey_id2]

report_cols = [
    "{id}",
    "Status",
    "Started on",
    "Geschlecht",
    "Altersgruppe",
    # "Höchster Bildungsabschluss",
    # "Beruf",
    "completion_time",
    "Bei dieser Frage ignorieren Sie bitte die folgenden Optionen und wählen Sie 'Stimme nicht zu'.",
    "valid",
]

# a function that translates ascii regex to python regex


def strip_df(df):
    """strip leading and trailing whitespaces from all columns and headers"""
    df = df.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
    df.columns = df.columns.str.strip()
    return df


def strip_double_whitespaces(df):
    df.columns = df.columns.str.replace(r"\s+", " ", regex=True)
    return df


def value_counts(df, normalize=True):
    """Count values in colums"""
    counts = {}
    for c in df:
        counts[c] = df[c].value_counts(normalize=normalize).to_dict()
    return counts


def to_results(counts, categories, labels=None, fact=100):
    results = {}
    for question, data in counts.items():
        label = question
        if labels:
            label = labels.get(question) or question
        results[label] = [data.get(k, 0.0) * 100 for k in categories]
    return results


def get_df(source):
    """read csv from file or surveyhero api"""
    return pd.read_csv(
        source,
        index_col="ID",
        date_format="%d.%m.%Y %H:%M:%S",
        parse_dates=["Started on", "Last updated on"],
        skipinitialspace=True,
    )


def make_request(survey_id, user, password):
    """ "make request to surveyhero api and return dataframe"""
    # get data from surveyhero api
    # params = {"format": "csv", "status": "completed"}
    params = {"format": "csv"}
    url = api_url.format(survey_id=survey_id)
    print(f"requesting data from {url}")
    r = requests.get(url, params=params, auth=(user, password))
    print(r.status_code)
    print(f"API request status code: {r.status_code}")
    source = StringIO(
        r.content.decode("utf-8").replace(
            "Direct Air Capture (DAC)\n", "Direct Air Capture (DAC)"
        )
    )
    df = get_df(source)
    # set attention column to default if survey has no attention column
    for k, v in attention_col.items():
        if k not in df.columns:
            df[k] = v
    return df


def get_data(
    source=None,
    user=None,
    password=None,
    translate=False,
    drop=True,
    survey_ids=None,
    validate=False,
    drop_invalid=False,
    set_dependent=False,
    no_knowledge_to_neutral=False,
    invert_risk=True,
):
    """read dataframe from file or surveyhero api"""
    if survey_ids is None:
        survey_ids = all_surveys
    if source is not None:
        # read from csv
        df = get_df(source)
    else:
        if not isinstance(survey_ids, list):
            survey_ids = [survey_ids]
        print(f"survey_ids: {survey_ids}")
        df = pd.concat(
            [make_request(survey_id, user, password) for survey_id in survey_ids]
        )

    if "completion_time" not in df.columns:
        df["completion_time"] = df["Last updated on"] - df["Started on"]
    df = strip_df(df)
    df = strip_double_whitespaces(df)
    df = df.replace(cleaning_dict)

    if set_dependent is True:
        df = set_dependent_questions(df)

    if no_knowledge_to_neutral is True:
        df = set_no_knowledge_to_neutral(df)

    if validate is True:
        df["valid"] = valid(df)

    if translate is True:
        df = df.rename(columns=translation_columns)
        df = df.replace(translation_answers)

    if drop is True:
        df = df.drop(columns=drop_cols)

    df = strip_df(df)

    if drop_invalid is True:
        df = df[df.valid == "valid"]

    return df


def adjust_excel_column_width(df, sheet):
    from openpyxl.utils import get_column_letter

    # Adjust the column widths based on the content
    for i, col in enumerate(df.columns):
        # Calculate the maximum width for the column
        max_width = max(df[col].astype(str).map(len).max(), len(col))
        # Set the column width
        # worksheet.set_column(i+1, i+1, max_width)
        sheet.column_dimensions[get_column_letter(i + 2)].width = max_width
    return sheet


def report_to_excel(df, filename):
    """Auto adjust column width and save to Excel file"""
    # Create a pandas.ExcelWriter object
    # if engine == "openpyxl":

    from openpyxl.styles import PatternFill

    GENDER_AGE = "Gender age groups (valid)"
    VALIDATION = "Validation results"

    df = df.astype({"completion_time": str, "{id}": str})
    writer = pd.ExcelWriter(filename, engine="openpyxl")
    df = df[report_cols]
    df.to_excel(writer, index=True, sheet_name=VALIDATION)

    # Get the XlsxWriter workbook and worksheet objects
    sheet = writer.sheets[VALIDATION]

    # writer.close()
    adjust_excel_column_width(df, sheet)
    # workbook = load_workbook(filename=filename)

    colnames = ["valid", "Geschlecht"]
    # get indices of column names
    col_indices = {
        cell.value: n
        for n, cell in enumerate(list(sheet.rows)[0])
        if cell.value in colnames
    }

    # highlight invalid rows
    for row in sheet.iter_rows(min_row=1, max_row=None, min_col=None, max_col=None):
        if row[col_indices["valid"]].value != "valid":
            for cell in row:
                cell.fill = PatternFill(
                    start_color="FF001F", end_color="FF001F", fill_type="solid"
                )

    gender_age(df).to_excel(writer, sheet_name=GENDER_AGE)
    adjust_excel_column_width(df, writer.sheets[GENDER_AGE])
    # workbook.save(filename=filename)
    writer.close()


def ensure_floats(df, groups):
    """Ensure to convert grouped columns into floats"""
    for k, v in groups.items():
        for col in v:
            df[col] = df[col].astype(float)
    return df
